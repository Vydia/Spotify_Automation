import openpyxl as pyxl
import pandas as pd
import numpy as np
import datetime as dt

"""
Things to do: 
Figure out how to incorporate FX CCY Checks to make ingestion easier. Also - ask if Intl vs. Domestic really matters? 
Modify Code to focus on copying 3 Situations - Clean, USD and Intl under a single Function. 
Modify Code to handle Currency Checks when copying the data - Needs Yes - USD, Yes - Intl, and No
Modify Code to handle passing through Sheet Number to simplify code to handle copying data more effectively

4/3/19
Need to modify code to handle currency check while copying data however, developed 3 different functions (1 using
iteration, 1 using an array to store the rows and delete columns using a reverse index, and one using iter_rows). The 
first two work. Need to tweak the third one and figure out which one works best. Works with 100 Row and 5000 Row file 
Need to implement time functions just to track speed of execution. 
Still need to handle FX Checks while copying the data. Maybe merging FX_Cleanup_Array function into the function to see if that
speeds things up. Create Functions Copy_Range_2, Paste_Range_2, and CopyData_2 to test that functionality.   
Find someone who can work with me on unit testing. Cases - file type, FX Output doesn't work. 
"""


# Sets Workbook Name and Loads workbook
wb_Spotify = pyxl.Workbook()
wb_Spotify = pyxl.load_workbook('Spotify_Test_XLS_2.xlsx') #Cannot do CSV
wb_Spotify_Sheets = []  # Allows for collection of Sheet Names
ws_Active = wb_Spotify.active # Defines Active Sheet Function
# Sets Header Row Data
HeaderRow = ['Source', 'Geocuntry','Product','URI','UPC','EAN','ISRC','Type','Title','Artist','Composer','Album Name',
             'Streams','Label','Payable invoice','Invoice currency','Payable EUR','Revenue',"FX Toggle"]

# Creates all sheets needed for processing the original file
ws_Active.title = "Spotify Raw"
ws_Clean = wb_Spotify.create_sheet('Spotify Clean')
ws_USD = wb_Spotify.create_sheet('Spotify USD')
ws_Intl = wb_Spotify.create_sheet('Spotify Intl')

# Sets Sheet Name List for Active Sheet Management
for i in range (len(wb_Spotify.sheetnames)):
    wb_Spotify_Sheets.append(wb_Spotify.sheetnames[i])

# Sets Starting and Ending Rows and Columns
Starting_Row = 1
Ending_Row = ws_Active.max_row
Starting_Column = 1
Ending_Column = len(HeaderRow)
print(Ending_Column)

# Fx used to copy a range of data
def Copy_Range(Start_Col, Start_Row, End_Col, End_Row, sheet):
    # Holds a full Array of Columns and Rows to be copied
    rangeSelected = []


    # Loops through selected Rows
    for i in range(Start_Row, End_Row + 1, 1):

        # Holds an Array of all the data for a given row
        rowSelected = []

        for j in range(Start_Col, End_Col + 1, 1):

            # Handles appending the array to house row data
            rowSelected.append(sheet.cell(row=i, column=j).value)


        # Appends the Range Array to store individual rows as part of a bigger array
        rangeSelected.append(rowSelected)

    return rangeSelected

"""

        Code for Looping and selecting cells - TBD if it's good or bad. May tweak in combination with FX_Cleanup

        if FX_Toggle == 0:
            if sheet.cell(row=i, column=19).value == FX_Toggle:  # For FX Specific Rows
                print(sheet.cell(row=i, column=len(HeaderRow)).value)

                for j in range(Start_Col, End_Col + 1, 1):
                    # Handles appending the array to house row data
                    rowSelected.append(sheet.cell(row=i, column=j).value)

        elif FX_Toggle == 1:

            if sheet.cell(row=i, column=19).value == FX_Toggle:  # For FX Specific Rows
                print(sheet.cell(row=i, column=len(HeaderRow)).value)

                for j in range(Start_Col, End_Col + 1, 1):
                    # Handles appending the array to house row data
                    rowSelected.append(sheet.cell(row=i, column=j).value)

        elif FX_Toggle == 2: 

            if sheet.cell(row=i, column=19).value == FX_Toggle:  # For FX Specific Rows
                print(sheet.cell(row=i, column=len(HeaderRow)).value)

                for j in range(Start_Col, End_Col + 1, 1):
                    # Handles appending the array to house row data
                    rowSelected.append(sheet.cell(row=i, column=j).value)

        else: #For all Rows
"""


# Takes the data that was copied and pastes it into a specific location
def Paste_Range(Start_Col, Start_Row, End_Col, End_Row, Sheet_Transfered, Copied_Data):

    #
    Count_Row = 0

    for i in range(Start_Row, End_Row + 1, 1):
        Count_Col = 0

        for j in range(Start_Col, End_Col + 1, 1):

            Sheet_Transfered.cell(row=i, column=j).value = Copied_Data[Count_Row][Count_Col]
            Count_Col += 1

        Count_Row += 1
    return


print(ws_Active.max_row)

# Fx for handling the copy over
def Copy_Data(Start_Col, Start_Row, End_Col, End_Row, Input_Sheet, Output_Sheet):
    Selected_Range = Copy_Range(Start_Col, Start_Row, End_Col, End_Row, Input_Sheet)
    Paste_Range(Start_Col, Start_Row, End_Col, End_Row, Output_Sheet, Selected_Range)
    return

#Handles Amending the Clean Sheet to read the way it should
def Amend_Clean_Data(Sheet_Name, HeaderRow):
    Sheet_Name.delete_rows(1,2)
    Sheet_Name.insert_cols(1,1)
    Sheet_Name.insert_cols(8,1)

    print(dt.time())
    for i in range (0, len(HeaderRow)):
        print(HeaderRow[i])

    Count_Row = 1
    Count_Col = 0


    for col in Sheet_Name.iter_cols(min_row=2, max_col=1, max_row=Sheet_Name.max_row):
        for cell in col:
            cell.value = "Spotify"

    for col in Sheet_Name.iter_cols(min_row=2, min_col=8, max_row=Sheet_Name.max_row, max_col=8):
        for cell in col:
                cell.value = "sound_recording"


    while Count_Col < len(HeaderRow):
        Sheet_Name.cell(row = 1, column = Count_Col+1).value = HeaderRow[Count_Col]
        Count_Col +=1

    Count_Row = 2
    for row in Sheet_Name.iter_cols(min_row=2, min_col=(len(HeaderRow)), max_col=(len(HeaderRow)), max_row=Sheet_Name.max_row):
        for cell in row:

            if Sheet_Name.cell(row = Count_Row, column = 16).value == "USD":
                cell.value = 1
            else:
                cell.value = 2

            Count_Row +=1

    return


'''

    Replaced this code with the Iter_Rows/Iter_Cols Function. 
    
    while Count_Row < (Sheet_Name.max_row+1):

        if Count_Row == 1:

            while Count_Col < (len(HeaderRow)):
                Sheet_Name.cell(row = 1, column = (Count_Col+1)).value = HeaderRow[Count_Col]
                Count_Col += 1
        else:
            Sheet_Name.cell(row=Count_Row, column=1).value = "Source"
            Sheet_Name.cell(row=Count_Row, column=8).value = "sound_recording"

        Count_Row +=1
    
    return
'''

def FX_Cleanup_Iter(FX_Toggle, sheet):
    # Trying to Automate using Iter_Row

    Count_Row = 2
    i = 1
    for row in sheet.iter_rows(min_row=2, min_col=(len(HeaderRow)), max_col=(len(HeaderRow)), max_row=sheet.max_row):
        for cell in row:

            if cell.value != FX_Toggle:
                sheet.delete_rows(Count_Row,1)
            else:
                Count_Row += 1

            print(i)
            i += 1
            print(Count_Row)



    return


def FX_Cleanup(FX_Toggle, sheet):

    #Slow Automation through all cells. Rewriting to try with Iter_Rows. This does work, but takes its sweet ass time
    #to go through a fuck ton of rows but it properly handles the deletion

    i = 2
    while i <= sheet.max_row:

        # iterates through and handles the incremental separately as to not screw with the index.

        if sheet.cell(row=i, column=len(HeaderRow)).value != FX_Toggle:
            sheet.delete_rows(i, 1)

        else:
            i += 1

        print(i)

    return


def FX_Cleanup_Array(FX_Toggle, Sheet):

    # Stores a list of all cells that are not included in teh FX_Toggle (ex. if USD = 1, EUR/GBP = 2)
    # Iterates through the array in reverse and deletes those cells to handle not messing up the index

    List_of_Cells = []

    for i in range (2, Sheet.max_row, 1):

        if Sheet.cell(row = i, column = 19).value != FX_Toggle:
            List_of_Cells.append(i)

    for i in range (len(List_of_Cells), 0, -1):
        Sheet.delete_rows(List_of_Cells[i-1],1)
        print(i)

    return



"""

Original FX_Cleanup Code. Simplified it a bit with the three versions above. Need to review for fastest option. 

def FX_Cleanup(CCY, sheet):

    if CCY == "USD":

        for i in range(2, sheet.max_row +1, 1):
            if sheet.cell(row = i+1, column = 16).value == "USD":
                continue
            else:
                sheet.delete_rows(i,1)
    else:
        for i in range(2, sheet.max_row +1, 1):
            if sheet.cell(row = i+1, column = 16).value != "USD":
                continue
            else:
                sheet.delete_rows(i,1)

    return

"""

print("Copying Raw Data.....")
Copy_Data(Starting_Column, Starting_Row, Ending_Column, Ending_Row, ws_Active, ws_Clean)
#print("Time Elapsed: ", dt.time()-Start_Time)
print("Cleaning Raw Data.....")
Amend_Clean_Data(ws_Clean, HeaderRow)
#print("Time Elapsed: ", dt.time()-Start_Time)
print("Copying USD Data.....")
Copy_Data(Starting_Column, Starting_Row, Ending_Column, Ending_Row, ws_Clean, ws_USD)
#print("Time Elapsed: ", dt.time()-Start_Time)
print("Copying Non-USD Data.....")
Copy_Data(Starting_Column, Starting_Row, Ending_Column, Ending_Row, ws_Clean, ws_Intl)
#print("Time Elapsed: ", dt.time()-Start_Time)

FX_Toggle += 1
print("Adjusting for USD FX")
FX_Cleanup_Array(FX_Toggle, ws_USD)
print(dt.time())
FX_Toggle +=1
print("Adjusting for Non-USD FX")
FX_Cleanup_Array(FX_Toggle, ws_Intl)
print(dt.time())
print("Saving Output in Excel Workbook...")
wb_Spotify.save('Spotify_Output_File.xlsx')
print(ws_Clean['A1'].value)
print(ws_Clean.cell(row=1,column=16).value)
print("Operation Complete")
#print("Total Time Elapsed: ", dt.time())